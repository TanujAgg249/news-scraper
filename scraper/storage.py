"""
storage.py - Data Storage, Cleanup & Visual Dashboard
=======================================================

This module handles:
1. Loading existing articles from the Excel file
2. Deduplicating new articles (by headline OR URL)
3. Auto-deleting articles older than MAX_ARTICLE_AGE_HOURS
4. Saving the combined data to a styled Excel file
5. Creating a "Dashboard" sheet with live Brent Crude price
6. Exporting a backup CSV
"""

import os
from datetime import datetime, timedelta, timezone

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from scraper.config import EXCEL_FILE, CSV_BACKUP_FILE, MAX_ARTICLE_AGE_HOURS


# ── Colour & Style Definitions ──────────────────────────────────

# Header styling
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Alternating row colours for readability
ROW_FILL_EVEN = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
ROW_FILL_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Cell styling
CELL_FONT = Font(size=10, name="Calibri")
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)

# Oil impact colour fills
BULLISH_FILL = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")  # Light green
BEARISH_FILL = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")  # Light red
NEUTRAL_FILL = PatternFill(start_color="FEF7E0", end_color="FEF7E0", fill_type="solid")  # Light yellow/orange

# Dashboard styles
DASH_BG = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")      # Dark navy
DASH_ACCENT = PatternFill(start_color="1B263B", end_color="1B263B", fill_type="solid")    # Slightly lighter
DASH_HIGHLIGHT = PatternFill(start_color="415A77", end_color="415A77", fill_type="solid") # Mid blue
DASH_TITLE_FONT = Font(color="E0E1DD", bold=True, size=14, name="Calibri")
DASH_LABEL_FONT = Font(color="778DA9", size=11, name="Calibri")
DASH_PRICE_FONT = Font(color="FFFFFF", bold=True, size=36, name="Calibri")
DASH_PRICE_FONT_UP = Font(color="00FF88", bold=True, size=36, name="Calibri")
DASH_PRICE_FONT_DOWN = Font(color="FF4444", bold=True, size=36, name="Calibri")
DASH_CHANGE_FONT_UP = Font(color="00FF88", bold=True, size=16, name="Calibri")
DASH_CHANGE_FONT_DOWN = Font(color="FF4444", bold=True, size=16, name="Calibri")
DASH_CHANGE_FONT_FLAT = Font(color="778DA9", bold=True, size=16, name="Calibri")
DASH_STAT_VALUE_FONT = Font(color="FFFFFF", bold=True, size=14, name="Calibri")
DASH_STAT_LABEL_FONT = Font(color="778DA9", size=10, name="Calibri")
DASH_TIMESTAMP_FONT = Font(color="778DA9", italic=True, size=9, name="Calibri")
DASH_CENTER = Alignment(horizontal="center", vertical="center")


def _load_existing() -> pd.DataFrame:
    """
    Load the existing Excel file into a DataFrame.
    If the file doesn't exist yet, return an empty DataFrame with the right columns.
    """
    if os.path.exists(EXCEL_FILE):
        try:
            df = pd.read_excel(EXCEL_FILE, sheet_name="Articles", engine="openpyxl")
            print(f"📂 Loaded {len(df)} existing records from {EXCEL_FILE}")
            return df
        except Exception as e:
            print(f"⚠️  Warning: Could not read {EXCEL_FILE}: {e}")
            print("   → Starting with empty data.")

    # Return empty DataFrame with expected columns
    return pd.DataFrame(columns=[
        "headline", "description", "source", "published_at", "url",
        "fetched_at", "matched_keywords", "oil_impact", "impact_reason"
    ])


def _cleanup_old_articles(df: pd.DataFrame) -> pd.DataFrame:
    """
    Remove articles older than MAX_ARTICLE_AGE_HOURS from the DataFrame.
    Returns the cleaned DataFrame.
    """
    if df.empty:
        return df

    try:
        # Parse published_at to datetime
        df["_pub_dt"] = pd.to_datetime(df["published_at"], errors="coerce", utc=True)

        cutoff = datetime.now(timezone.utc) - timedelta(hours=MAX_ARTICLE_AGE_HOURS)
        before_count = len(df)

        # Keep only articles newer than the cutoff (and those we couldn't parse)
        df = df[(df["_pub_dt"] >= cutoff) | (df["_pub_dt"].isna())]

        removed = before_count - len(df)
        if removed > 0:
            print(f"🗑️  Removed {removed} article(s) older than {MAX_ARTICLE_AGE_HOURS} hours")

        # Drop the temp column
        df = df.drop(columns=["_pub_dt"])

    except Exception as e:
        print(f"⚠️  Warning: Cleanup failed: {e}")

    return df.reset_index(drop=True)


def save_to_excel(articles: list[dict], oil_price_data: dict | None = None) -> int:
    """
    Save new articles to the Excel file, skipping duplicates.
    Removes articles older than MAX_ARTICLE_AGE_HOURS.
    Creates a Dashboard sheet with Brent Crude price.

    Args:
        articles: List of article dicts from fetcher.py
        oil_price_data: Optional dict from oil_price.get_brent_crude_price()

    Returns:
        The number of NEW (non-duplicate) articles added.
    """

    # Step 1: Load existing data
    existing_df = _load_existing()

    # Step 2: Clean up old articles from existing data (before merging new ones)
    existing_df = _cleanup_old_articles(existing_df)

    # Step 3: Convert new articles to a DataFrame
    new_df = pd.DataFrame(articles)

    if existing_df.empty:
        # No existing data — everything is new
        combined_df = new_df
        new_count = len(new_df)
    else:
        # Step 4: Find duplicates by checking headline AND url
        existing_headlines = set(existing_df["headline"].dropna().str.strip().str.lower())
        existing_urls = set(existing_df["url"].dropna().str.strip().str.lower())

        # Keep only articles where BOTH headline and url are new
        mask = new_df.apply(
            lambda row: (
                row["headline"].strip().lower() not in existing_headlines
                and row["url"].strip().lower() not in existing_urls
            ),
            axis=1,
        )
        unique_new = new_df[mask]
        new_count = len(unique_new)

        # Combine old + new
        combined_df = pd.concat([existing_df, unique_new], ignore_index=True)

    # Step 5: Sort by publication time (newest first)
    combined_df = combined_df.sort_values(by="published_at", ascending=False)
    combined_df = combined_df.reset_index(drop=True)

    # Step 6: Write to Excel with styling
    try:
        # Write data to "Articles" sheet
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
            combined_df.to_excel(writer, index=False, sheet_name="Articles")
    except PermissionError:
        print(f"❌ ERROR: Cannot write to {EXCEL_FILE} — is it open in another app?")
        print("   → Close the file and try again.")
        return 0
    except Exception as e:
        print(f"❌ ERROR: Failed to save Excel file: {e}")
        return 0

    # Step 7: Apply visual styling + dashboard
    _apply_styling(combined_df, oil_price_data)

    print(f"📊 Total records in {EXCEL_FILE}: {len(combined_df)}")
    return new_count


def _apply_styling(df: pd.DataFrame, oil_price_data: dict | None = None):
    """
    Open the saved Excel file and apply:
    1. Professional styling to the Articles sheet
    2. A Dashboard sheet with Brent Crude price
    """
    try:
        wb = load_workbook(EXCEL_FILE)

        # ── Style the Articles sheet ─────────────────────────
        ws = wb["Articles"]

        # Style the header row
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        # Find special column indices
        headers = [cell.value for cell in ws[1]]
        url_col_idx = None
        impact_col_idx = None
        if "url" in headers:
            url_col_idx = headers.index("url") + 1  # 1-indexed
        if "oil_impact" in headers:
            impact_col_idx = headers.index("oil_impact") + 1

        # Style data rows with alternating colours (+ oil impact color-coding)
        for row_num in range(2, ws.max_row + 1):
            # Check if this row has an oil_impact value for color-coding
            row_fill = ROW_FILL_EVEN if row_num % 2 == 0 else ROW_FILL_ODD
            if impact_col_idx:
                impact_value = str(ws.cell(row=row_num, column=impact_col_idx).value or "").strip()
                if impact_value == "Bullish":
                    row_fill = BULLISH_FILL
                elif impact_value == "Bearish":
                    row_fill = BEARISH_FILL
                elif impact_value == "Neutral":
                    row_fill = NEUTRAL_FILL

            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.fill = row_fill
                cell.font = CELL_FONT
                cell.alignment = CELL_ALIGNMENT
                cell.border = THIN_BORDER

                # Make URLs clickable hyperlinks
                if col_num == url_col_idx and cell.value:
                    url = str(cell.value).strip()
                    if url.startswith("http"):
                        cell.hyperlink = url
                        cell.font = Font(
                            size=10, name="Calibri",
                            color="1155CC", underline="single"
                        )
                        cell.value = url

        # Set column widths
        column_widths = {
            "headline": 60,
            "description": 45,
            "source": 18,
            "published_at": 22,
            "url": 40,
            "fetched_at": 20,
            "matched_keywords": 25,
            "oil_impact": 15,
            "impact_reason": 50,
        }

        headers = [cell.value for cell in ws[1]]
        for i, header in enumerate(headers):
            col_letter = get_column_letter(i + 1)
            width = column_widths.get(header, 15)
            ws.column_dimensions[col_letter].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # ── Create the Dashboard sheet ───────────────────────
        if "Dashboard" in wb.sheetnames:
            del wb["Dashboard"]

        ws_dash = wb.create_sheet("Dashboard", 0)  # Insert as first sheet
        _build_dashboard(ws_dash, oil_price_data, len(df))

        wb.save(EXCEL_FILE)
        print("🎨 Styling and dashboard applied.")

    except Exception as e:
        print(f"⚠️  Warning: Could not apply styling: {e}")


def _build_dashboard(ws, oil_price_data: dict | None, article_count: int):
    """
    Build a visually appealing dashboard sheet with:
    - Live Brent Crude price in large font
    - Price change indicator
    - Article count summary
    """

    # Set dark background for the entire visible area
    for row in range(1, 25):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.fill = DASH_BG

    # Set column widths
    ws.column_dimensions["A"].width = 3     # Left margin
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 3     # Right margin
    ws.column_dimensions["G"].width = 3
    ws.column_dimensions["H"].width = 3
    ws.column_dimensions["I"].width = 3

    # Set row heights
    ws.row_dimensions[1].height = 15    # Top margin
    ws.row_dimensions[2].height = 25    # Title
    ws.row_dimensions[3].height = 10    # Spacer
    ws.row_dimensions[4].height = 20    # Label
    ws.row_dimensions[5].height = 55    # Big price
    ws.row_dimensions[6].height = 30    # Change
    ws.row_dimensions[7].height = 10    # Spacer
    ws.row_dimensions[8].height = 15    # Divider line
    ws.row_dimensions[9].height = 10    # Spacer
    ws.row_dimensions[10].height = 25   # Stats title
    ws.row_dimensions[11].height = 35   # Stat values
    ws.row_dimensions[12].height = 20   # Stat labels

    # ── Title row ──
    title_cell = ws.cell(row=2, column=2, value="🛢️  BRENT CRUDE OIL — LIVE DASHBOARD")
    title_cell.font = DASH_TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("B2:E2")

    if oil_price_data:
        price = oil_price_data["price"]
        change = oil_price_data["change"]
        change_pct = oil_price_data["change_pct"]
        timestamp = oil_price_data["timestamp"]

        # Determine up/down/flat
        if change > 0:
            price_font = DASH_PRICE_FONT_UP
            change_font = DASH_CHANGE_FONT_UP
            arrow = "▲"
        elif change < 0:
            price_font = DASH_PRICE_FONT_DOWN
            change_font = DASH_CHANGE_FONT_DOWN
            arrow = "▼"
        else:
            price_font = DASH_PRICE_FONT
            change_font = DASH_CHANGE_FONT_FLAT
            arrow = "●"

        # ── "Current Price" label ──
        label_cell = ws.cell(row=4, column=2, value="CURRENT PRICE (USD)")
        label_cell.font = DASH_LABEL_FONT
        label_cell.alignment = Alignment(horizontal="left", vertical="bottom")

        # ── Big price display ──
        price_cell = ws.cell(row=5, column=2, value=f"${price:.2f}")
        price_cell.font = price_font
        price_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells("B5:C5")

        # ── Change display ──
        change_text = f"{arrow} {change:+.2f} ({change_pct:+.2f}%)"
        change_cell = ws.cell(row=6, column=2, value=change_text)
        change_cell.font = change_font
        change_cell.alignment = Alignment(horizontal="left", vertical="top")
        ws.merge_cells("B6:C6")

        # ── Accent bar for price area ──
        for row in range(4, 7):
            for col in range(2, 4):
                ws.cell(row=row, column=col).fill = DASH_ACCENT

        # ── Divider ──
        for col in range(2, 6):
            divider_cell = ws.cell(row=8, column=col)
            divider_cell.fill = DASH_HIGHLIGHT
            divider_cell.border = Border(
                bottom=Side(style="thin", color="415A77")
            )

        # ── Stats section ──
        stats_title = ws.cell(row=10, column=2, value="📊 SESSION STATS")
        stats_title.font = Font(color="E0E1DD", bold=True, size=12, name="Calibri")
        stats_title.alignment = Alignment(horizontal="left", vertical="center")

        # Stat: Article count
        ws.cell(row=11, column=2, value=str(article_count)).font = DASH_STAT_VALUE_FONT
        ws.cell(row=11, column=2).alignment = DASH_CENTER
        ws.cell(row=11, column=2).fill = DASH_ACCENT
        ws.cell(row=12, column=2, value="Articles Today").font = DASH_STAT_LABEL_FONT
        ws.cell(row=12, column=2).alignment = DASH_CENTER

        # Stat: Price
        ws.cell(row=11, column=3, value=f"${price:.2f}").font = DASH_STAT_VALUE_FONT
        ws.cell(row=11, column=3).alignment = DASH_CENTER
        ws.cell(row=11, column=3).fill = DASH_ACCENT
        ws.cell(row=12, column=3, value="Brent Crude").font = DASH_STAT_LABEL_FONT
        ws.cell(row=12, column=3).alignment = DASH_CENTER

        # Stat: Change
        ws.cell(row=11, column=4, value=f"{change_pct:+.2f}%").font = change_font
        ws.cell(row=11, column=4).alignment = DASH_CENTER
        ws.cell(row=11, column=4).fill = DASH_ACCENT
        ws.cell(row=12, column=4, value="Daily Change").font = DASH_STAT_LABEL_FONT
        ws.cell(row=12, column=4).alignment = DASH_CENTER

        # Stat: Last updated
        ws.cell(row=11, column=5, value=timestamp).font = DASH_STAT_VALUE_FONT
        ws.cell(row=11, column=5).alignment = DASH_CENTER
        ws.cell(row=11, column=5).fill = DASH_ACCENT
        ws.cell(row=12, column=5, value="Last Updated").font = DASH_STAT_LABEL_FONT
        ws.cell(row=12, column=5).alignment = DASH_CENTER

        # ── Timestamp footer ──
        ts_cell = ws.cell(row=14, column=2, value=f"Data refreshed: {timestamp}")
        ts_cell.font = DASH_TIMESTAMP_FONT
        ts_cell.alignment = Alignment(horizontal="left", vertical="center")

    else:
        # No price data available
        na_cell = ws.cell(row=5, column=2, value="Price unavailable")
        na_cell.font = Font(color="FF4444", bold=True, size=20, name="Calibri")
        na_cell.alignment = Alignment(horizontal="left", vertical="center")

        hint_cell = ws.cell(row=6, column=2, value="Could not fetch Brent Crude price. Check internet connection.")
        hint_cell.font = DASH_LABEL_FONT
        hint_cell.alignment = Alignment(horizontal="left", vertical="top")
        ws.merge_cells("B6:E6")

    # Hide gridlines for cleaner look
    ws.sheet_view.showGridLines = False


def export_backup_csv():
    """
    Export the current Excel data as a CSV backup.
    This provides a second copy in a universal format.
    """
    if not os.path.exists(EXCEL_FILE):
        print("⚠️  No Excel file to backup yet.")
        return

    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name="Articles", engine="openpyxl")
        df.to_csv(CSV_BACKUP_FILE, index=False)
        print(f"📁 CSV backup saved: {CSV_BACKUP_FILE} ({len(df)} records)")
    except Exception as e:
        print(f"⚠️  Warning: CSV backup failed: {e}")
